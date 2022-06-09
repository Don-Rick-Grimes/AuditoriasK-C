from cProfile import label
from functools import partial
import os
from openpyxl import load_workbook
import numpy as np
from tkinter import Button, Label, Tk, filedialog, Entry

root = Tk()
root.title('Factura Auditoria')
root.geometry('450x450+300+200')

asignacion: int
label1 = Label(root, font=("Arial", 15))
label2 = Label(root, font=("Arial", 14))
boton1 = Button(root, fg='green', font=("Arial", 25))
boton2 = Button(root, font=("Arial", 25))
boton3 = Button(root, font=("Arial", 25))

def abrir():
    archivo = filedialog.askopenfilename(title="abrir", filetypes=(('Libro Excel (*.xlsx)','*.xlsx'),))
    wb = load_workbook(archivo)

    ws_datos = wb[wb.sheetnames[1]]

    cell_range = ws_datos['A':'B']
    factura = [[],[]]
    llaves = []
    ventas = []
    for l,v in zip(cell_range[0][:], cell_range[1][:]):
        llaves.append(l.value)
        ventas.append(v.value)

    indice_datos = 0
    venta = 0

    for llave in range(min(llaves),max(llaves)+1):
        while indice_datos <=(len(llaves)-1) and llaves[indice_datos] == llave:
            venta = venta + ventas[indice_datos]
            indice_datos = indice_datos + 1

        if llaves[indice_datos-1] == llave:
            factura[0].append(llave)
            factura[1].append(venta)
            venta = 0

    for i in range(0,len(factura[1])):
        if factura[1][i] > asignacion:
            factura[1][i] = asignacion

    total_factura =sum(factura[1])

    a = np.array(factura)
    a = a.transpose()
    factura = a.tolist()

    wb.create_sheet('FACTURA')
    ws_factura = wb['FACTURA']

    ws_factura.append(['#LLAVE', 'CONSUMO'])
    for r in factura:
        ws_factura.append(r)

    ws_factura.append([None, total_factura])

    label2['text'] = f'Total factura: {total_factura} pesos'
    label2['fg'] = 'black'
    boton1['text'] = 'Guardar Excel'
    boton1['command'] = partial(guardar,wb) 


def guardar(wb):
    archivo = filedialog.asksaveasfilename(title="guardar", filetypes=(('Libro Excel (*.xlsx)','*.xlsx'),))
    wb.save(archivo+'.xlsx')
    boton1['text'] = 'Calcular otra factura'
    boton1['command'] = nuevaFactura
    boton2['text'] = 'Salir'
    boton2['command'] = close
    boton2['fg'] = 'blue'
    boton2.pack()


def go(valorAsignado: int):
    boton2.pack_forget()
    boton3.pack_forget()
    label1['text'] = f"Asignacion: {valorAsignado}"
    global asignacion
    asignacion = valorAsignado
    label2['text'] = "En la 2da hoja del excel deben estar los registros\n de las ventas en la forma LLAVE/VALOR."
    label2['fg'] = 'goldenrod'
    boton1['text'] = 'Abrir Excel'
    boton1['command'] = abrir


def close():
   root.quit()


def otroValor():
    pass

def nuevaFactura():
    label1['text']= "Seleccione el valor de la asignacion:"
    label1.pack()
    label2['text'] = ''
    label2.pack()
    boton1['text'] = '21500'
    boton1['command'] = partial(go,21500)
    boton1.pack()
    boton2['text'] = '55000'
    boton2['fg'] = 'green'
    boton2['command'] = partial(go,55000)
    boton2.pack()
    boton3['text'] = 'Otro Valor'
    boton3['fg'] = 'green'
    boton3['command'] = otroValor
    boton3.pack()

nuevaFactura()
root.mainloop()
